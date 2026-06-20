import pandas as pd
import os
import openpyxl as oxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from system import log
import fnmatch

PREFIXIES = ["C", "R", "D", "F", "X", "S", "Q", "VT", "VD"]

def set_column_autowidth(ws, columns, reserve=1.2):
    """
    Устанавливает оптимальную ширину столбцов на основе содержимого.
    """
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Получаем букву столбца (A, B, C, ...)
        
        if column in columns:
            # Находим максимальную длину текста в столбце
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
            # Устанавливаем ширину столбца с небольшим запасом
            adjusted_width = (max_length + 2) * reserve  # Можно изменить коэффициент для более комфортного отображения
            ws.column_dimensions[column].width = adjusted_width

def get_max_split(line: str, references):
    max = -1
    result = ""
    for ref in references:
        if max < len(line.split(ref)):
            result = line.split(ref)
            max = len(line.split(ref))
    return result

def checkPrefix(partnumber: str, exceptPrefix: bool):
    pos = partnumber.find("-")
    if pos != -1:
        if exceptPrefix and (partnumber[0:pos+1] in [s + "-" for s in PREFIXIES]):
            return partnumber[pos+1:].strip()
        else:
            return partnumber.strip()
    else:
        return partnumber

def is_number(s: str):
    """
    Проверяет, является ли строка числом типа int
    """
    try:
        int(s)
        return True
    except ValueError:
        return False

def findNumber(line: str):
    """
    Ищет целое число в строке
    """
    for i in range(0,len(line)):
        for j in range(len(line), 0, -1):
            if is_number(line[i:j]):
                return(str(line[:i]), int(line[i:j]))
    return -1

def generatePnp(table: pd.Dataframe, type_file_name, inputType, outputType, outputName, exceptPrefix, reper_type, reper_list, reper_match, openFile=True):
    
    file = ""
    components = {}
    hand = []
    dnp = []
    try:
        for index, row in table.iterrows():
            if row['ref'] != "PCB":
                references = get_max_split(row['ref'], [",", " "])
                references_flat = []

                for ref in references:
                    if "-" in ref:
                        r, startr = findNumber(ref[:ref.find("-")])
                        r_, endr = findNumber(ref[ref.find("-")+1:])
                        
                        for r_n in range(startr, endr+1):
                            references_flat.append(f"{r}{r_n}")
                    else:
                        references_flat.append(ref)

                for ref_flat in references_flat:
                    partNumber = str(row["pn"])
                    partNumber = checkPrefix(partNumber, exceptPrefix)
                    components[ref_flat] = partNumber
                    if row['tm'] == 'HAND':
                        hand.append(ref_flat)
                    elif row['tm'] == 'NM' or row['tm'] == 'DNP' or row['tm'] == 'NOT MOUNT':
                        dnp.append(ref_flat)
    except Exception as e:
        log(f"Неизвестная ошибка: {e}\nКонтекст: specification file")
        raise
    try:
        match inputType:
            case "Файл place (Allegro)":
                place = pd.read_table(type_file_name, skiprows=1, header=None, encoding='windows-1251')
                place = place.sort_values(by=0)

                output = []
                match reper_type:
                    case 0:
                        for line in reper_list:
                            if line.strip()[0] != "#":
                                line_s = line.split(" ")
                                match line_s[1].lower():
                                    case "top":
                                        layer = "TopLayer"
                                    case "bottom":
                                        layer = "BottomLayer"
                                    case _:
                                        log(f"Реперная точка {line_s[0]}: не указан слой или он некорректный.")
                                output.append([line_s[0], "REPER", layer, f"{float(line_s[2]):.3f}", f"{float(line_s[3]):.3f}", f"{float(line_s[4]):.3f}"])
                    case 1:
                        pass

                for index, row in place.iterrows():
                    line = row.values[0]
                    split_line = line.split()
                    ref = split_line[0]
                    #partNumber = checkPrefix(partNumber, exceptPrefix)
                    if ref in components.keys():
                        if len(split_line) == 5:
                            output.append([ref, components[ref], "TopLayer", split_line[1], split_line[2], split_line[3]])
                        else:
                            output.append([ref, components[ref], "BottomLayer", split_line[1], split_line[2], split_line[3]])
                    
                output_df = pd.DataFrame(output)

            case "Файл PnP SiDeCo":

                place = pd.read_excel(type_file_name, skiprows=1, header=None, engine='xlrd')
                place[1] = place[1].astype(str)
                print(place)
                place = place.sort_values(by=1)
                output = []

                match reper_type:
                    case 0:
                        for line in reper_list:
                            if line.strip()[0] != "#":
                                line_s = line.split(" ")
                                match line_s[1].lower():
                                    case "top":
                                        layer = "TopLayer"
                                    case "bottom":
                                        layer = "BottomLayer"
                                    case _:
                                        log(f"Реперная точка {line_s[0]}: не указан слой или он некорректный.")
                                output.append([line_s[0], "REPER", layer, f"{float(line_s[2]):.3f}", f"{float(line_s[3]):.3f}", f"{float(line_s[4]):.3f}"])
                    case 1:
                        for index, row in place.iterrows():
                            ref = str(row[1])
                            if fnmatch.fnmatch(ref, f"{reper_match}"):
                                if row[4] == "Top":
                                    output.append([ref, "REPER", "TopLayer", row[6], row[7], row[5]])
                                else:
                                    output.append([ref, "REPER", "BottomLayer", row[6], row[7], row[5]])            

                for index, row in place.iterrows():

                    ref = str(row[1])
                    if ref in components.keys():
                        if row[4] == "Top":
                            output.append([ref, components[ref], "TopLayer", row[6], row[7], row[5]])
                        else:
                            output.append([ref, components[ref], "BottomLayer", row[6], row[7], row[5]])
                output_df = pd.DataFrame(output)
            case _:
                pass
    except TypeError as e:
        log(f"Ошибка типа данных: {e}\nВ одном из столбцов неверный тип данных (либо NaN)\nКонтекст: {inputType}")
        raise
    except Exception as e:
        log(f"Неизвестная ошибка: {e}\nКонтекст: {inputType}")
        raise
            

        

    match outputType:
        case "Стандартный":

            file = f"{outputName}_PnP.txt"
            output_df.to_csv(file, sep='\t', header=None, index=None)
        case "Файл .xlsx с цветной маркировкой":

            file = f"{outputName}_PnP.xlsx"
            
            name = outputName[outputName.rfind("/")+1:]
            output_df[2] = output_df[2].replace({'TopLayer': 'TOP', 'BottomLayer': 'BOTTOM'})

            output_df.to_excel(file, header=["REFDES", "PART_NUMBER", "LAYER", "LocationX", "LocationY", "Rotation"], index=None, startrow=4)

            workbook = oxl.load_workbook(file)
            sheet = workbook.worksheets[0]
            sheet.title = f"PnP"

            color_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            color_blue = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            color_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

            sheet.cell(row=1, column=9).value = "SMT Mount"
            sheet.cell(row=1, column=9).fill = color_green
            sheet.cell(row=2, column=9).value = "DIP or Manually Mount"
            sheet.cell(row=2, column=9).fill = color_blue
            sheet.cell(row=3, column=9).value = "Not Mount"
            sheet.cell(row=3, column=9).fill = color_yellow

            row_n = 6
            for index, row in output_df.iterrows():
                ref = row[0]
                if row[1] != "REPER":
                    if ref in hand:
                        color = color_blue
                    elif ref in dnp:
                        color = color_yellow
                    else:
                        color = color_green
                    for column in range(1,len(row)+1):
                        sheet.cell(row=row_n+index, column=column).fill = color
                
            set_column_autowidth(sheet, ["A", "B", "C", "D", "E", "F", "G", "H", "I"])

            sheet.cell(row=1, column=1).value = f"{name}"
            sheet.cell(row=2, column=1).value = "Report Origin = (0.000)"
            sheet.cell(row=3, column=1).value = "Units used = mm"

            workbook.save(file)
            

        case _:
            pass

    if openFile and file != "":
        os.startfile(file)



if __name__ == "__main__":

    # from compare.tools import get_table

    # first_file_name = "CBoard_TEST_V2_RAPTOR_SP.xlsx"
    # first_file_count_column = "B"
    # first_file_pn_column = "D"
    # first_file_skip_row = 8
    # first_file_mount = "F"

    # table = get_table(first_file_name, f"{first_file_count_column}, C, {first_file_pn_column}, {first_file_mount}, L", first_file_skip_row, ["count", "ref", "pn", "tm", "dpn"])

    # if 0:
    #     table = table[table['tm'] != 'DNP']
    #     table = table[table['tm'] != 'NM']
        
    # pass

    pos = "J10-12"

    print(pos[-1:2])

    print(findNumber(pos[:pos.find("-")]))

    print(findNumber(pos[pos.find("-")+1:]))
